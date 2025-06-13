# + Nombre de archivo: test_database.py
# + Ubicación de archivo: tests/test_database.py
# User-provided custom instructions
import os
import sys
import importlib
from pathlib import Path
from datetime import datetime
import pytest
import openpyxl

sqlalchemy = pytest.importorskip("sqlalchemy")
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

# Agregar ruta del paquete
ROOT_DIR = Path(__file__).resolve().parents[1]
sys.path.append(str(ROOT_DIR / "Sandy bot"))

import tests.telegram_stub  # Registra las clases fake de telegram

# Stub de dotenv requerido por config
dotenv_stub = importlib.util.module_from_spec(
    importlib.machinery.ModuleSpec("dotenv", None)
)
dotenv_stub.load_dotenv = lambda *a, **k: None
sys.modules.setdefault("dotenv", dotenv_stub)

# Variables de entorno necesarias para Config
required_vars = {
    "TELEGRAM_TOKEN": "x",
    "OPENAI_API_KEY": "x",
    "NOTION_TOKEN": "x",
    "NOTION_DATABASE_ID": "x",
    "DB_USER": "user",
    "DB_PASSWORD": "pass",
    "SLACK_WEBHOOK_URL": "x",
    "SUPERVISOR_DB_ID": "x",
    "DB_HOST": "localhost",
    "DB_PORT": "5432",
    "DB_NAME": "sandy",
}
os.environ.update(required_vars)

# Forzar que ``sandybot.database`` utilice SQLite en memoria
import sqlalchemy

orig_create_engine = sqlalchemy.create_engine
sqlalchemy.create_engine = lambda *a, **k: orig_create_engine("sqlite:///:memory:")

bd = importlib.import_module("sandybot.database")

sqlalchemy.create_engine = orig_create_engine
bd.SessionLocal = sessionmaker(bind=bd.engine, expire_on_commit=False)
bd.Base.metadata.create_all(bind=bd.engine)


def test_buscar_servicios_por_camara():
    bd.crear_servicio(nombre="S1", cliente="A", camaras=["Cámara Central"])
    bd.crear_servicio(nombre="S2", cliente="B", camaras=["Nodo Secundario"])
    bd.crear_servicio(nombre="S3", cliente="C", camaras=["Avenida General San Martin"])

    res1 = bd.buscar_servicios_por_camara("camara central")
    assert {s.nombre for s in res1} == {"S1"}

    # Búsqueda exacta encerrando el nombre entre comillas
    res_exact = bd.buscar_servicios_por_camara("camara central", exacto=True)
    assert {s.nombre for s in res_exact} == {"S1"}

    # "central" solo debería coincidir si no se exige exactitud
    res_no = bd.buscar_servicios_por_camara("central", exacto=True)
    assert res_no == []

    res2 = bd.buscar_servicios_por_camara("gral. san martin")
    assert {s.nombre for s in res2} == {"S3"}

    # Caso con abreviaturas y acentos que antes causaba falso negativo
    camara = "Cra Av. Gral Juan Domingo Per\u00f3n 7540 BENAVIDEZ"
    bd.crear_servicio(nombre="S4", cliente="D", camaras=[camara])

    # La búsqueda debería funcionar aunque se omitan los acentos
    res3 = bd.buscar_servicios_por_camara("peron 7540")
    assert {s.nombre for s in res3} == {"S4"}

    bd.crear_servicio(nombre="S5", cliente="E", camaras=["Cámara Fiscalía"])
    res4 = bd.buscar_servicios_por_camara("camara fiscalia")
    assert {s.nombre for s in res4} == {"S5"}


def test_buscar_servicios_por_camara_jsonb():
    """Verifica la búsqueda cuando ``camaras`` se almacena como JSONB."""
    bd.crear_servicio(nombre="SJ1", cliente="G", camaras=["Cámara JSONB"])

    res = bd.buscar_servicios_por_camara("camara jsonb")
    assert {s.nombre for s in res} == {"SJ1"}

    res_exact = bd.buscar_servicios_por_camara("camara jsonb", exacto=True)
    assert {s.nombre for s in res_exact} == {"SJ1"}


def test_exportar_camaras_servicio(tmp_path):
    servicio = bd.crear_servicio(
        nombre="S4", cliente="D", camaras=["Camara 1", "Camara 2"]
    )

    ruta = tmp_path / "camaras.xlsx"
    ok = bd.exportar_camaras_servicio(servicio.id, str(ruta))
    assert ok is True
    assert ruta.exists()

    wb = openpyxl.load_workbook(ruta)
    ws = wb.active
    filas = [c[0].value for c in ws.iter_rows(values_only=False)]
    assert filas == ["camara", "Camara 1", "Camara 2"]


def test_exportar_camaras_servicio_cadena(tmp_path):
    """Verifica la exportación cuando las cámaras se guardaron como texto JSON."""
    servicio = bd.crear_servicio(nombre="S4b", cliente="D", camaras='["C1", "C2"]')

    ruta = tmp_path / "camaras_str.xlsx"
    ok = bd.exportar_camaras_servicio(servicio.id, str(ruta))
    assert ok is True
    assert ruta.exists()

    wb = openpyxl.load_workbook(ruta)
    ws = wb.active
    filas = [c[0].value for c in ws.iter_rows(values_only=False)]
    assert filas == ["camara", "C1", "C2"]


def test_actualizar_tracking_jsonb():
    servicio = bd.crear_servicio(nombre="S6", cliente="F")
    bd.actualizar_tracking(
        servicio.id, "ruta.txt", ["C1"], ["t1.txt"], tipo="principal"
    )

    with bd.SessionLocal() as s:
        reg = s.get(bd.Servicio, servicio.id)
        assert reg.ruta_tracking == "ruta.txt"
        assert reg.camaras == ["C1"]
        assert isinstance(reg.trackings[0], dict)
        assert reg.trackings[0]["ruta"] == "t1.txt"
        assert reg.trackings[0]["tipo"] == "principal"


def test_actualizar_tracking_string():
    """Verifica que se actualice si el campo ``trackings`` quedó como texto."""
    servicio = bd.crear_servicio(nombre="S7", cliente="G", trackings="[]")
    bd.actualizar_tracking(
        servicio.id, trackings_txt=["nuevo.txt"], tipo="complementario"
    )

    with bd.SessionLocal() as s:
        reg = s.get(bd.Servicio, servicio.id)
        assert reg.trackings[0]["ruta"] == "nuevo.txt"
        assert reg.trackings[0]["tipo"] == "complementario"


def test_crear_ingreso():
    servicio = bd.crear_servicio(nombre="S5", cliente="E")
    fecha = datetime(2023, 1, 1, 12, 30)
    ingreso = bd.crear_ingreso(servicio.id, "Camara X", fecha=fecha, usuario="u")
    with bd.SessionLocal() as session:
        fila = session.query(bd.Ingreso).first()
        assert fila.camara == "Camara X"
        assert fila.fecha == fecha


def test_registrar_servicio_merge():
    """Verifica que ``registrar_servicio`` no duplique filas."""
    bd.crear_servicio(id=100, nombre="n1", cliente="a")
    bd.registrar_servicio(100, "c1")
    bd.registrar_servicio(100, "c1")
    with bd.SessionLocal() as session:
        filas = session.query(bd.Servicio).filter(bd.Servicio.id == 100).all()
        assert len(filas) == 1
        assert filas[0].id_carrier == "c1"


def test_ensure_servicio_columns_crea_indice():
    """Verifica que ``ensure_servicio_columns`` genere el índice."""

    with bd.engine.begin() as conn:
        conn.execute(text("DROP INDEX IF EXISTS ix_servicios_id_carrier"))

    insp = sqlalchemy.inspect(bd.engine)
    assert not any(
        i["name"] == "ix_servicios_id_carrier" for i in insp.get_indexes("servicios")
    )

    bd.ensure_servicio_columns()

    insp = sqlalchemy.inspect(bd.engine)
    assert any(
        i["name"] == "ix_servicios_id_carrier" for i in insp.get_indexes("servicios")
    )


def test_ensure_servicio_columns_cliente():
    """La función crea la tabla de clientes y la columna cliente_id."""

    with bd.engine.begin() as conn:
        conn.execute(text("DROP TABLE IF EXISTS clientes"))

    insp = sqlalchemy.inspect(bd.engine)
    assert "clientes" not in insp.get_table_names()

    bd.ensure_servicio_columns()

    insp = sqlalchemy.inspect(bd.engine)
    assert "clientes" in insp.get_table_names()
    cols = {c["name"] for c in insp.get_columns("servicios")}
    assert "cliente_id" in cols


def test_cliente_destinatarios():
    """Los servicios se vinculan con un cliente y sus correos."""

    with bd.SessionLocal() as s:
        cli = bd.Cliente(
            nombre="Acme",
            destinatarios=["a@x.com"],
            destinatarios_carrier={"Telco": ["b@x.com"]},
        )
        s.add(cli)
        s.commit()
        s.refresh(cli)
        servicio = bd.crear_servicio(
            nombre="S_cli", cliente="Acme", cliente_id=cli.id, carrier="Telco"
        )

        dest = bd.obtener_destinatarios_servicio(servicio.id)
        assert dest == ["b@x.com"]


def test_crear_tarea_y_relacion():
    s = bd.crear_servicio(nombre="Srv", cliente="Cli")
    tarea = bd.crear_tarea_programada(
        datetime(2024, 1, 1, 8),
        datetime(2024, 1, 1, 10),
        "Mantenimiento",
        [s.id],
        tiempo_afectacion="2h",
    )
    tareas = bd.obtener_tareas_servicio(s.id)
    assert len(tareas) == 1
    assert tareas[0].id == tarea.id


def test_carrier_asociaciones():
    car = bd.Carrier(nombre="CarrierTest")
    with bd.SessionLocal() as s:
        s.add(car)
        s.commit()
        s.refresh(car)

    srv = bd.crear_servicio(nombre="SrvC", cliente="CliC", carrier_id=car.id)
    tarea = bd.crear_tarea_programada(
        datetime(2024, 1, 3, 8),
        datetime(2024, 1, 3, 10),
        "Mant",
        [srv.id],
        carrier_id=car.id,
    )

    with bd.SessionLocal() as s:
        s_srv = s.get(bd.Servicio, srv.id)
        s_tarea = s.get(bd.TareaProgramada, tarea.id)

    assert s_srv.carrier_id == car.id
    assert s_tarea.carrier_id == car.id


def test_ensure_servicio_columns_indice_tarea_programada():
    """La función crea el índice combinado de fechas en tareas_programadas."""
    # 1️⃣  El índice se elimina si existe para garantizar la prueba
    with bd.engine.begin() as conn:
        conn.execute(
            text("DROP INDEX IF EXISTS ix_tareas_programadas_fecha_inicio_fecha_fin")
        )

    insp = sqlalchemy.inspect(bd.engine)
    assert not any(
        i["name"] == "ix_tareas_programadas_fecha_inicio_fecha_fin"
        for i in insp.get_indexes("tareas_programadas")
    )

    # 2️⃣  Se vuelve a invocar la función; debe crear el índice
    bd.ensure_servicio_columns()

    insp = sqlalchemy.inspect(bd.engine)
    assert any(
        i["name"] == "ix_tareas_programadas_fecha_inicio_fecha_fin"
        for i in insp.get_indexes("tareas_programadas")
    )


def test_tarea_servicio_unica():
    """La relación tarea-servicio no debe duplicarse (restricción única)."""
    s = bd.crear_servicio(nombre="SrvU", cliente="CliU")
    tarea = bd.crear_tarea_programada(
        datetime(2024, 1, 2, 8),
        datetime(2024, 1, 2, 10),
        "Mantenimiento",
        [s.id],
    )

    # 1️⃣  El primer insert manual debe violar la restricción UNIQUE
    with pytest.raises(sqlalchemy.exc.IntegrityError):
        with bd.SessionLocal() as session:
            session.add(bd.TareaServicio(tarea_id=tarea.id, servicio_id=s.id))
            session.commit()


def test_crear_tarea_varios_servicios():
    s1 = bd.crear_servicio(nombre="Sv1", cliente="C1")
    s2 = bd.crear_servicio(nombre="Sv2", cliente="C2")
    tarea = bd.crear_tarea_programada(
        datetime(2024, 1, 3, 8),
        datetime(2024, 1, 3, 10),
        "Mantenimiento",
        [s1.id, s2.id],
    )
    with bd.SessionLocal() as session:
        rels = (
            session.query(bd.TareaServicio)
            .filter(bd.TareaServicio.tarea_id == tarea.id)
            .all()
        )
    assert len(rels) == 2
    assert {r.servicio_id for r in rels} == {s1.id, s2.id}


def test_crear_tarea_servicio_repetido():
    """Crear una tarea con el mismo servicio dos veces provoca un error."""
    s = bd.crear_servicio(nombre="SvRep", cliente="CliR")

    with pytest.raises(sqlalchemy.exc.IntegrityError):
        bd.crear_tarea_programada(
            datetime(2024, 1, 4, 8),
            datetime(2024, 1, 4, 10),
            "Mantenimiento",
            [s.id, s.id],
        )


def test_reclamos_por_servicio():
    srv1 = bd.crear_servicio(nombre="SrvRec1", cliente="Cli")
    srv2 = bd.crear_servicio(nombre="SrvRec2", cliente="Cli")
    fecha = datetime(2024, 5, 1, 12)
    cierre = datetime(2024, 5, 2, 14)
    bd.crear_reclamo(
        srv1.id,
        "R1",
        fecha_inicio=fecha,
        fecha_cierre=cierre,
        tipo_solucion="TS",
        descripcion_solucion="Sol",
    )
    bd.crear_reclamo(srv2.id, "R2")

    recs1 = bd.obtener_reclamos_servicio(srv1.id)
    recs2 = bd.obtener_reclamos_servicio(srv2.id)

    assert len(recs1) == 1
    assert recs1[0].numero == "R1"
    assert recs1[0].fecha_inicio == fecha
    assert recs1[0].fecha_cierre == cierre
    assert recs1[0].tipo_solucion == "TS"
    assert recs1[0].descripcion_solucion == "Sol"
    assert len(recs2) == 1
    assert recs2[0].numero == "R2"


def test_reclamo_unico():
    """Reclamo con el mismo número no debe duplicarse."""
    srv = bd.crear_servicio(nombre="SrvDup", cliente="Cli")
    r1 = bd.crear_reclamo(srv.id, "DUP")
    r2 = bd.crear_reclamo(srv.id, "DUP")
    with bd.SessionLocal() as s:
        filas = (
            s.query(bd.Reclamo)
            .filter(bd.Reclamo.servicio_id == srv.id)
            .all()
        )
    assert len(filas) == 1
    assert r1.id == r2.id


def test_camara_unica():
    """Cámara repetida para un servicio retorna el mismo registro."""
    srv = bd.crear_servicio(nombre="SrvCam", cliente="Cli")
    c1 = bd.crear_camara("Cam", srv.id)
    c2 = bd.crear_camara("Cam", srv.id)
    with bd.SessionLocal() as s:
        filas = (
            s.query(bd.Camara)
            .filter(bd.Camara.id_servicio == srv.id)
            .all()
        )
    assert len(filas) == 1
    assert c1.id == c2.id
